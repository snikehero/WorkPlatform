import io
import json
import os
import re
import uuid
from datetime import date, datetime, timedelta, timezone
from enum import Enum
from urllib.parse import quote

from fastapi import Depends, FastAPI, File, Form, HTTPException, UploadFile, status
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import StreamingResponse
from fastapi.security import OAuth2PasswordBearer
from jose import JWTError, jwt
from openpyxl import load_workbook
from passlib.context import CryptContext
from pydantic import BaseModel, Field
from sqlalchemy import Boolean, Date, DateTime, ForeignKey, Integer, String, Text, create_engine, select, text
from sqlalchemy.orm import DeclarativeBase, Mapped, Session, mapped_column, relationship, sessionmaker


DATABASE_URL = os.getenv("DATABASE_URL", "postgresql+psycopg://postgres:postgres@postgres:5432/workplatform")
JWT_SECRET = os.getenv("JWT_SECRET", "change-me")
JWT_ALGORITHM = "HS256"
JWT_EXP_MINUTES = int(os.getenv("JWT_EXP_MINUTES", "720"))


class Base(DeclarativeBase):
    pass


class UserRole(str, Enum):
    admin = "admin"
    developer = "developer"
    user = "user"


class User(Base):
    __tablename__ = "users"

    id: Mapped[str] = mapped_column(String(36), primary_key=True, default=lambda: str(uuid.uuid4()))
    email: Mapped[str] = mapped_column(String(320), unique=True, index=True)
    password_hash: Mapped[str] = mapped_column(String(255))
    role: Mapped[str] = mapped_column(String(20), default=UserRole.user.value)
    preferred_language: Mapped[str] = mapped_column(String(5), default="en")
    created_at: Mapped[datetime] = mapped_column(DateTime(timezone=True), default=lambda: datetime.now(timezone.utc))


class Project(Base):
    __tablename__ = "projects"

    id: Mapped[str] = mapped_column(String(36), primary_key=True, default=lambda: str(uuid.uuid4()))
    owner_id: Mapped[str] = mapped_column(String(36), ForeignKey("users.id"), index=True)
    name: Mapped[str] = mapped_column(String(200))
    description: Mapped[str] = mapped_column(Text, default="")
    created_at: Mapped[datetime] = mapped_column(DateTime(timezone=True), default=lambda: datetime.now(timezone.utc))


class Task(Base):
    __tablename__ = "tasks"

    id: Mapped[str] = mapped_column(String(36), primary_key=True, default=lambda: str(uuid.uuid4()))
    owner_id: Mapped[str] = mapped_column(String(36), ForeignKey("users.id"), index=True)
    title: Mapped[str] = mapped_column(String(250))
    details: Mapped[str] = mapped_column(Text, default="")
    status: Mapped[str] = mapped_column(String(20), default="todo")
    project_id: Mapped[str | None] = mapped_column(String(36), ForeignKey("projects.id"), nullable=True)
    task_date: Mapped[date] = mapped_column(Date)
    created_at: Mapped[datetime] = mapped_column(DateTime(timezone=True), default=lambda: datetime.now(timezone.utc))


class Note(Base):
    __tablename__ = "notes"

    id: Mapped[str] = mapped_column(String(36), primary_key=True, default=lambda: str(uuid.uuid4()))
    owner_id: Mapped[str] = mapped_column(String(36), ForeignKey("users.id"), index=True)
    title: Mapped[str] = mapped_column(String(250))
    content: Mapped[str] = mapped_column(Text, default="")
    note_date: Mapped[date] = mapped_column(Date)
    created_at: Mapped[datetime] = mapped_column(DateTime(timezone=True), default=lambda: datetime.now(timezone.utc))


class TeamEvent(Base):
    __tablename__ = "team_events"

    id: Mapped[str] = mapped_column(String(36), primary_key=True, default=lambda: str(uuid.uuid4()))
    owner_id: Mapped[str] = mapped_column(String(36), ForeignKey("users.id"), index=True)
    title: Mapped[str] = mapped_column(String(250))
    event_date: Mapped[date] = mapped_column(Date)
    description: Mapped[str] = mapped_column(Text, default="")
    owner: Mapped[str] = mapped_column(String(200), default="")
    location: Mapped[str] = mapped_column(String(200), default="")
    created_at: Mapped[datetime] = mapped_column(DateTime(timezone=True), default=lambda: datetime.now(timezone.utc))


class Ticket(Base):
    __tablename__ = "tickets"

    id: Mapped[str] = mapped_column(String(36), primary_key=True, default=lambda: str(uuid.uuid4()))
    requester_id: Mapped[str] = mapped_column(String(36), ForeignKey("users.id"), index=True)
    title: Mapped[str] = mapped_column(String(250))
    description: Mapped[str] = mapped_column(Text, default="")
    category: Mapped[str] = mapped_column(String(50), default="help")
    priority: Mapped[str] = mapped_column(String(20), default="medium")
    status: Mapped[str] = mapped_column(String(20), default="new")
    resolution: Mapped[str] = mapped_column(Text, default="")
    process_notes: Mapped[str] = mapped_column(Text, default="")
    evidence_json: Mapped[str] = mapped_column(Text, default="[]")
    fixed_by_id: Mapped[str | None] = mapped_column(String(36), ForeignKey("users.id"), nullable=True)
    assignee_id: Mapped[str | None] = mapped_column(String(36), ForeignKey("users.id"), nullable=True, index=True)
    first_response_due_at: Mapped[datetime | None] = mapped_column(DateTime(timezone=True), nullable=True)
    resolution_due_at: Mapped[datetime | None] = mapped_column(DateTime(timezone=True), nullable=True)
    first_responded_at: Mapped[datetime | None] = mapped_column(DateTime(timezone=True), nullable=True)
    resolved_at: Mapped[datetime | None] = mapped_column(DateTime(timezone=True), nullable=True)
    closed_at: Mapped[datetime | None] = mapped_column(DateTime(timezone=True), nullable=True)
    created_at: Mapped[datetime] = mapped_column(DateTime(timezone=True), default=lambda: datetime.now(timezone.utc))
    updated_at: Mapped[datetime] = mapped_column(DateTime(timezone=True), default=lambda: datetime.now(timezone.utc))


class TicketEvent(Base):
    __tablename__ = "ticket_events"

    id: Mapped[str] = mapped_column(String(36), primary_key=True, default=lambda: str(uuid.uuid4()))
    ticket_id: Mapped[str] = mapped_column(String(36), ForeignKey("tickets.id"), index=True)
    actor_id: Mapped[str | None] = mapped_column(String(36), ForeignKey("users.id"), nullable=True, index=True)
    event_type: Mapped[str] = mapped_column(String(50))
    payload_json: Mapped[str] = mapped_column(Text, default="{}")
    created_at: Mapped[datetime] = mapped_column(DateTime(timezone=True), default=lambda: datetime.now(timezone.utc), index=True)


class MaintenanceRecord(Base):
    __tablename__ = "maintenance_records"

    id: Mapped[str] = mapped_column(String(36), primary_key=True, default=lambda: str(uuid.uuid4()))
    owner_id: Mapped[str] = mapped_column(String(36), ForeignKey("users.id"), index=True)
    maintenance_date: Mapped[date] = mapped_column(Date)
    qr: Mapped[str] = mapped_column(String(120))
    brand: Mapped[str] = mapped_column(String(120))
    model: Mapped[str] = mapped_column(String(200))
    user_name: Mapped[str] = mapped_column(String(250))
    serial_number: Mapped[str] = mapped_column(String(250))
    consecutive: Mapped[str] = mapped_column(String(50))
    maintenance_type: Mapped[str] = mapped_column(String(1))
    location: Mapped[str] = mapped_column(String(250))
    responsible_name: Mapped[str] = mapped_column(String(250))
    created_at: Mapped[datetime] = mapped_column(DateTime(timezone=True), default=lambda: datetime.now(timezone.utc))
    checks: Mapped[list["MaintenanceCheck"]] = relationship(
        "MaintenanceCheck",
        back_populates="record",
        cascade="all, delete-orphan",
    )


class MaintenanceCheck(Base):
    __tablename__ = "maintenance_checks"

    id: Mapped[str] = mapped_column(String(120), primary_key=True)
    record_id: Mapped[str] = mapped_column(String(36), ForeignKey("maintenance_records.id"), primary_key=True)
    label: Mapped[str] = mapped_column(Text)
    category: Mapped[str] = mapped_column(String(20))
    checked: Mapped[bool] = mapped_column(default=False)
    observation: Mapped[str] = mapped_column(Text, default="")
    record: Mapped[MaintenanceRecord] = relationship("MaintenanceRecord", back_populates="checks")


class Notification(Base):
    __tablename__ = "notifications"

    id: Mapped[str] = mapped_column(String(36), primary_key=True, default=lambda: str(uuid.uuid4()))
    owner_id: Mapped[str] = mapped_column(String(36), ForeignKey("users.id"), index=True)
    title: Mapped[str] = mapped_column(String(250))
    message: Mapped[str] = mapped_column(Text, default="")
    category: Mapped[str] = mapped_column(String(30), default="info")
    due_date: Mapped[date | None] = mapped_column(Date, nullable=True)
    is_read: Mapped[bool] = mapped_column(Boolean, default=False)
    created_at: Mapped[datetime] = mapped_column(DateTime(timezone=True), default=lambda: datetime.now(timezone.utc))


class KnowledgeArticle(Base):
    __tablename__ = "knowledge_articles"

    id: Mapped[str] = mapped_column(String(36), primary_key=True, default=lambda: str(uuid.uuid4()))
    owner_id: Mapped[str] = mapped_column(String(36), ForeignKey("users.id"), index=True)
    title: Mapped[str] = mapped_column(String(250))
    summary: Mapped[str] = mapped_column(Text, default="")
    content: Mapped[str] = mapped_column(Text, default="")
    tags: Mapped[str] = mapped_column(Text, default="")
    created_at: Mapped[datetime] = mapped_column(DateTime(timezone=True), default=lambda: datetime.now(timezone.utc))
    updated_at: Mapped[datetime] = mapped_column(DateTime(timezone=True), default=lambda: datetime.now(timezone.utc))


class Asset(Base):
    __tablename__ = "assets"

    id: Mapped[str] = mapped_column(String(36), primary_key=True, default=lambda: str(uuid.uuid4()))
    owner_id: Mapped[str] = mapped_column(String(36), ForeignKey("users.id"), index=True)
    asset_tag: Mapped[str] = mapped_column(String(120))
    # Legacy field kept for backward-compatible inserts/updates on existing databases.
    name: Mapped[str] = mapped_column(String(250))
    qr_code: Mapped[str] = mapped_column(String(250), default="")
    location: Mapped[str] = mapped_column(String(250), default="")
    serial_number: Mapped[str] = mapped_column(String(250), default="")
    category: Mapped[str] = mapped_column(String(80), default="")
    manufacturer: Mapped[str] = mapped_column(String(120), default="")
    model: Mapped[str] = mapped_column(String(120), default="")
    supplier: Mapped[str] = mapped_column(String(120), default="")
    status: Mapped[str] = mapped_column(String(30), default="active")
    # Legacy fields kept for backward-compatible inserts/updates on existing databases.
    assigned_to: Mapped[str] = mapped_column(String(250), default="")
    purchase_date: Mapped[date | None] = mapped_column(Date, nullable=True)
    warranty_until: Mapped[date | None] = mapped_column(Date, nullable=True)
    notes: Mapped[str] = mapped_column(Text, default="")
    user_name: Mapped[str] = mapped_column(String(250), default="")
    condition: Mapped[str] = mapped_column(String(120), default="")
    created_at: Mapped[datetime] = mapped_column(DateTime(timezone=True), default=lambda: datetime.now(timezone.utc))
    updated_at: Mapped[datetime] = mapped_column(DateTime(timezone=True), default=lambda: datetime.now(timezone.utc))


class Person(Base):
    __tablename__ = "people"

    id: Mapped[str] = mapped_column(String(36), primary_key=True, default=lambda: str(uuid.uuid4()))
    user_id: Mapped[str | None] = mapped_column(String(36), ForeignKey("users.id"), nullable=True, unique=True)
    legacy_id: Mapped[int | None] = mapped_column(Integer, nullable=True, unique=True)
    type: Mapped[str] = mapped_column(String(10), default="user")
    role_id: Mapped[int] = mapped_column(Integer, default=2)
    client_id: Mapped[int] = mapped_column(Integer, default=1)
    name: Mapped[str] = mapped_column(String(128))
    email: Mapped[str] = mapped_column(String(128), default="")
    title: Mapped[str] = mapped_column(String(128), default="")
    role: Mapped[str] = mapped_column(String(128), default="")
    job_department: Mapped[str] = mapped_column(String(128), default="")
    mobile: Mapped[str] = mapped_column(String(100), default="")
    password_hash_legacy: Mapped[str] = mapped_column(String(128), default="")
    theme: Mapped[str] = mapped_column(String(64), default="skin-blue")
    sidebar: Mapped[str] = mapped_column(String(64), default="opened")
    layout: Mapped[str] = mapped_column(String(64), default="")
    notes: Mapped[str] = mapped_column(Text, default="")
    signature: Mapped[str] = mapped_column(Text, default="")
    session_id: Mapped[str] = mapped_column(String(255), default="")
    reset_key: Mapped[str] = mapped_column(String(255), default="")
    auto_refresh: Mapped[int] = mapped_column(Integer, default=0)
    lang: Mapped[str] = mapped_column(String(5), default="en")
    tickets_notification: Mapped[bool] = mapped_column(Boolean, default=False)
    avatar_info: Mapped[str] = mapped_column(Text, default="")
    created_at: Mapped[datetime] = mapped_column(DateTime(timezone=True), default=lambda: datetime.now(timezone.utc))
    updated_at: Mapped[datetime] = mapped_column(DateTime(timezone=True), default=lambda: datetime.now(timezone.utc))


engine = create_engine(DATABASE_URL, pool_pre_ping=True)
SessionLocal = sessionmaker(bind=engine, autoflush=False, autocommit=False)
pwd_context = CryptContext(schemes=["pbkdf2_sha256"], deprecated="auto")
oauth2_scheme = OAuth2PasswordBearer(tokenUrl="/api/auth/login")


def get_db():
    db = SessionLocal()
    try:
        yield db
    finally:
        db.close()


def hash_password(password: str) -> str:
    return pwd_context.hash(password)


def verify_password(plain: str, hashed: str) -> bool:
    return pwd_context.verify(plain, hashed)


def create_access_token(user: User) -> str:
    expire = datetime.now(timezone.utc) + timedelta(minutes=JWT_EXP_MINUTES)
    payload = {"sub": user.id, "email": user.email, "role": user.role, "exp": expire}
    return jwt.encode(payload, JWT_SECRET, algorithm=JWT_ALGORITHM)


def get_current_user(
    token: str = Depends(oauth2_scheme),
    db: Session = Depends(get_db),
) -> User:
    try:
        payload = jwt.decode(token, JWT_SECRET, algorithms=[JWT_ALGORITHM])
        user_id = str(payload.get("sub"))
    except JWTError as exc:
        raise HTTPException(status_code=status.HTTP_401_UNAUTHORIZED, detail="Invalid token") from exc

    user = db.get(User, user_id)
    if not user:
        raise HTTPException(status_code=status.HTTP_401_UNAUTHORIZED, detail="Invalid user")
    return user


class AuthRegisterIn(BaseModel):
    email: str
    password: str = Field(min_length=6, max_length=128)
    role: UserRole = UserRole.user


class AuthLoginIn(BaseModel):
    email: str
    password: str


class AuthOut(BaseModel):
    token: str
    user_email: str
    role: str
    preferred_language: str


class AdminUserOut(BaseModel):
    id: str
    email: str
    role: str
    preferredLanguage: str
    createdAt: str


class AdminCreateUserIn(BaseModel):
    email: str
    password: str = Field(min_length=6, max_length=128)
    role: UserRole = UserRole.user


class AdminUpdateUserIn(BaseModel):
    email: str
    role: UserRole


class AdminResetPasswordIn(BaseModel):
    password: str = Field(min_length=6, max_length=128)


class PersonIn(BaseModel):
    name: str
    email: str = ""
    title: str = ""
    role: UserRole = UserRole.user
    department: str = ""
    mobile: str = ""
    notes: str = ""


class PersonOut(BaseModel):
    id: str
    userId: str | None
    userEmail: str | None
    legacyId: int | None
    name: str
    email: str
    title: str
    role: str
    department: str
    mobile: str
    notes: str
    createdAt: str
    updatedAt: str


class ProjectIn(BaseModel):
    name: str
    description: str = ""


class ProjectOut(BaseModel):
    id: str
    name: str
    description: str
    createdAt: str


class TaskIn(BaseModel):
    title: str
    details: str = ""
    projectId: str | None = None
    taskDate: str


class TaskStatusPatch(BaseModel):
    status: str


class TaskOut(BaseModel):
    id: str
    title: str
    details: str
    status: str
    projectId: str | None
    taskDate: str
    createdAt: str


class NoteIn(BaseModel):
    title: str
    content: str = ""
    noteDate: str


class NoteOut(BaseModel):
    id: str
    title: str
    content: str
    noteDate: str
    createdAt: str


class TeamEventIn(BaseModel):
    title: str
    eventDate: str
    description: str = ""
    owner: str = ""
    location: str = ""


class TeamEventOut(BaseModel):
    id: str
    title: str
    eventDate: str
    description: str
    owner: str
    location: str
    createdAt: str


class TicketIn(BaseModel):
    title: str
    description: str = ""
    category: str = "help"
    priority: str = "medium"


class TicketEvidenceIn(BaseModel):
    id: str
    text: str = ""
    imageData: str | None = None
    imageName: str | None = None
    createdAt: str | None = None


class TicketEvidenceOut(BaseModel):
    id: str
    text: str
    imageData: str | None
    imageName: str | None
    createdAt: str


class TicketPatchIn(BaseModel):
    status: str | None = None
    assigneeId: str | None = None
    resolution: str = ""
    processNotes: str | None = None
    evidence: list[TicketEvidenceIn] | None = None


class TicketAssignIn(BaseModel):
    assigneeId: str | None = None


class TicketOut(BaseModel):
    id: str
    requesterId: str
    requesterEmail: str
    title: str
    description: str
    category: str
    priority: str
    status: str
    resolution: str
    processNotes: str
    evidence: list[TicketEvidenceOut]
    assigneeId: str | None
    assigneeEmail: str | None
    slaState: str
    firstResponseDueAt: str | None
    resolutionDueAt: str | None
    firstRespondedAt: str | None
    resolvedAt: str | None
    closedAt: str | None
    fixedById: str | None
    fixedByEmail: str | None
    createdAt: str
    updatedAt: str


class TicketEventOut(BaseModel):
    id: str
    ticketId: str
    actorId: str | None
    actorEmail: str | None
    eventType: str
    payload: dict[str, object]
    createdAt: str


class TicketAssigneeOut(BaseModel):
    id: str
    email: str
    role: str


class MaintenanceCheckPayload(BaseModel):
    id: str
    label: str
    category: str
    checked: bool
    observation: str = ""


class MaintenanceRecordIn(BaseModel):
    maintenanceDate: str
    qr: str
    brand: str
    model: str
    user: str
    serialNumber: str
    consecutive: str
    maintenanceType: str
    location: str
    responsibleName: str
    checks: list[MaintenanceCheckPayload]


class MaintenanceRecordOut(BaseModel):
    id: str
    maintenanceDate: str
    qr: str
    brand: str
    model: str
    user: str
    serialNumber: str
    consecutive: str
    maintenanceType: str
    location: str
    responsibleName: str
    checks: list[MaintenanceCheckPayload]
    createdAt: str


class NotificationIn(BaseModel):
    title: str
    message: str = ""
    category: str = "info"
    dueDate: str | None = None


class NotificationReadPatchIn(BaseModel):
    read: bool


class NotificationOut(BaseModel):
    id: str
    title: str
    message: str
    category: str
    dueDate: str | None
    read: bool
    createdAt: str


class KnowledgeArticleIn(BaseModel):
    title: str
    summary: str = ""
    content: str = ""
    tags: list[str] = []


class KnowledgeArticleOut(BaseModel):
    id: str
    title: str
    summary: str
    content: str
    tags: list[str]
    createdAt: str
    updatedAt: str


class AssetIn(BaseModel):
    assetTag: str = ""
    qrCode: str = ""
    qrClass: str = "A"
    location: str = ""
    serialNumber: str = ""
    category: str = ""
    manufacturer: str = ""
    model: str = ""
    supplier: str = ""
    status: str = "active"
    user: str = ""
    condition: str = ""
    notes: str = ""


class AssetOut(BaseModel):
    id: str
    assetTag: str
    qrCode: str
    location: str
    serialNumber: str
    category: str
    manufacturer: str
    model: str
    supplier: str
    status: str
    user: str
    condition: str
    notes: str
    createdAt: str
    updatedAt: str


class ChangePasswordIn(BaseModel):
    currentPassword: str
    newPassword: str = Field(min_length=6, max_length=128)


class LanguagePreferenceIn(BaseModel):
    preferredLanguage: str


UNASSIGNED_USER_LABEL = "Unassigned"


def parse_date(value: str) -> date:
    try:
        return datetime.strptime(value, "%Y-%m-%d").date()
    except ValueError as exc:
        raise HTTPException(status_code=400, detail="Date must be YYYY-MM-DD") from exc


def parse_optional_date(value: str | None) -> date | None:
    if value is None:
        return None
    raw = value.strip()
    if not raw:
        return None
    return parse_date(raw)


def parse_tags(raw: str) -> list[str]:
    if not raw.strip():
        return []
    return [item for item in (segment.strip() for segment in raw.split(",")) if item]


def serialize_tags(items: list[str]) -> str:
    normalized: list[str] = []
    seen: set[str] = set()
    for item in items:
        value = item.strip()
        if not value:
            continue
        lowered = value.lower()
        if lowered in seen:
            continue
        seen.add(lowered)
        normalized.append(value)
    return ",".join(normalized)


DEPARTMENT_OPTIONS: tuple[str, ...] = (
    "Direccion",
    "Administracion",
    "Comercial",
    "Capital Humano",
    "Calidad",
    "Compras",
    "Automatizacion",
    "Diseño Electrico",
    "HMI",
    "BMS",
    "Mantenimiento",
    "Construccion",
    "Desarrollo",
    "Electricidad y Fuerza",
    "Administracion de Proyectos",
    "Past Employee",
)

TICKET_STATUS_VALUES: tuple[str, ...] = (
    "new",
    "triaged",
    "in_progress",
    "waiting_user",
    "blocked",
    "resolved",
    "closed",
    "reopened",
)
TICKET_ACTIVE_STATUSES: tuple[str, ...] = ("new", "triaged", "in_progress", "waiting_user", "blocked", "reopened")
TICKET_RESOLUTION_STATUSES: tuple[str, ...] = ("resolved", "closed")
TICKET_TRANSITIONS: dict[str, set[str]] = {
    "new": {"triaged", "in_progress", "blocked", "resolved"},
    "triaged": {"in_progress", "waiting_user", "blocked", "resolved"},
    "in_progress": {"waiting_user", "blocked", "resolved"},
    "waiting_user": {"in_progress", "blocked", "resolved"},
    "blocked": {"triaged", "in_progress", "waiting_user", "resolved"},
    "resolved": {"closed", "reopened"},
    "closed": {"reopened"},
    "reopened": {"triaged", "in_progress", "waiting_user", "blocked", "resolved"},
}
TICKET_PRIORITY_VALUES: tuple[str, ...] = ("low", "medium", "high", "critical")
TICKET_CATEGORY_VALUES: tuple[str, ...] = ("printer", "help", "network", "software", "hardware", "access")
TICKET_FIRST_RESPONSE_SLA_HOURS: dict[str, int] = {"low": 24, "medium": 8, "high": 2, "critical": 1}
TICKET_RESOLUTION_SLA_HOURS: dict[str, int] = {"low": 72, "medium": 24, "high": 8, "critical": 4}


def normalize_department(value: str) -> str:
    raw = (value or "").strip()
    if not raw:
        return DEPARTMENT_OPTIONS[0]
    normalized = raw.casefold()
    for option in DEPARTMENT_OPTIONS:
        if option.casefold() == normalized:
            return option
    allowed = "|".join(DEPARTMENT_OPTIONS)
    raise HTTPException(status_code=400, detail=f"department must be one of: {allowed}")


def build_next_asset_tag(db: Session) -> str:
    tags = db.scalars(select(Asset.asset_tag)).all()
    highest = 0
    pattern = re.compile(r"^TDC-(\d{4,})$", re.IGNORECASE)
    for tag in tags:
        match = pattern.match((tag or "").strip())
        if not match:
            continue
        value = int(match.group(1))
        if value > highest:
            highest = value
    return f"TDC-{highest + 1:04d}"


def normalize_qr_class(value: str | None) -> str:
    qr_class = (value or "A").strip().upper()
    if qr_class not in ("A", "B", "C"):
        raise HTTPException(status_code=400, detail="qrClass must be A|B|C")
    return qr_class


def build_qr_code_from_asset_tag(asset_tag: str, qr_class: str) -> str:
    tag_match = re.match(r"^TDC-(\d{4,})$", (asset_tag or "").strip(), flags=re.IGNORECASE)
    if not tag_match:
        raise HTTPException(status_code=400, detail="assetTag must match TDC-####")
    consecutive = int(tag_match.group(1))
    current_year = datetime.now(timezone.utc).year % 100
    year_segment = f"{current_year:02d}"
    return f"TDC-{year_segment}-{consecutive:04d}-{qr_class}"


def normalize_login_identity(value: str) -> str:
    raw = (value or "").strip().lower()
    if not raw:
        raise HTTPException(status_code=400, detail="Email/username is required")
    if "@" in raw:
        return raw
    return f"{raw}@workplatform.local"


def username_from_email(value: str) -> str:
    email = normalize_login_identity(value)
    return email.split("@", 1)[0].upper()


def to_iso(value: datetime) -> str:
    return value.astimezone(timezone.utc).isoformat()


def normalize_assigned_user(value: str | None) -> str:
    assigned = (value or "").strip()
    return assigned if assigned else UNASSIGNED_USER_LABEL


def normalize_ticket_category(value: str) -> str:
    category = (value or "").strip().lower()
    if category not in TICKET_CATEGORY_VALUES:
        allowed = "|".join(TICKET_CATEGORY_VALUES)
        raise HTTPException(status_code=400, detail=f"category must be {allowed}")
    return category


def normalize_ticket_priority(value: str) -> str:
    priority = (value or "").strip().lower()
    if priority not in TICKET_PRIORITY_VALUES:
        allowed = "|".join(TICKET_PRIORITY_VALUES)
        raise HTTPException(status_code=400, detail=f"priority must be {allowed}")
    return priority


def normalize_ticket_status(value: str) -> str:
    status_value = (value or "").strip().lower()
    if status_value not in TICKET_STATUS_VALUES:
        allowed = "|".join(TICKET_STATUS_VALUES)
        raise HTTPException(status_code=400, detail=f"status must be {allowed}")
    return status_value


def validate_ticket_transition(current: str, target: str) -> None:
    if current == target:
        return
    allowed = TICKET_TRANSITIONS.get(current, set())
    if target not in allowed:
        raise HTTPException(status_code=400, detail=f"invalid transition {current}->{target}")


def calculate_ticket_deadlines(priority: str, created_at: datetime) -> tuple[datetime, datetime]:
    first_due = created_at + timedelta(hours=TICKET_FIRST_RESPONSE_SLA_HOURS[priority])
    resolution_due = created_at + timedelta(hours=TICKET_RESOLUTION_SLA_HOURS[priority])
    return first_due, resolution_due


def compute_ticket_sla_state(ticket: Ticket) -> str:
    if ticket.status in TICKET_RESOLUTION_STATUSES:
        return "completed"
    now = datetime.now(timezone.utc)
    due = ticket.resolution_due_at
    if due is None:
        return "on_track"
    if due <= now:
        return "breached"
    if due <= now + timedelta(hours=2):
        return "at_risk"
    return "on_track"


def log_ticket_event(db: Session, ticket: Ticket, actor_id: str | None, event_type: str, payload: dict[str, object] | None = None) -> None:
    event = TicketEvent(
        ticket_id=ticket.id,
        actor_id=actor_id,
        event_type=event_type,
        payload_json=json.dumps(payload or {}),
    )
    db.add(event)


def validate_assignment_permission(current_user: User, assignee_id: str | None) -> None:
    if current_user.role == UserRole.admin.value:
        return
    if current_user.role == UserRole.developer.value:
        if assignee_id is None or assignee_id == current_user.id:
            return
        raise HTTPException(status_code=403, detail="Developers can only assign tickets to themselves")
    raise HTTPException(status_code=403, detail="Only admin or developer can assign tickets")


def require_admin(current_user: User = Depends(get_current_user)) -> User:
    if current_user.role != UserRole.admin.value:
        raise HTTPException(status_code=403, detail="Admin role required")
    return current_user


def require_developer_or_admin(current_user: User = Depends(get_current_user)) -> User:
    if current_user.role not in (UserRole.admin.value, UserRole.developer.value):
        raise HTTPException(status_code=403, detail="Developer or admin role required")
    return current_user


def project_to_out(project: Project) -> ProjectOut:
    return ProjectOut(id=project.id, name=project.name, description=project.description, createdAt=to_iso(project.created_at))


def task_to_out(task: Task) -> TaskOut:
    return TaskOut(
        id=task.id,
        title=task.title,
        details=task.details,
        status=task.status,
        projectId=task.project_id,
        taskDate=task.task_date.isoformat(),
        createdAt=to_iso(task.created_at),
    )


def note_to_out(note: Note) -> NoteOut:
    return NoteOut(
        id=note.id,
        title=note.title,
        content=note.content,
        noteDate=note.note_date.isoformat(),
        createdAt=to_iso(note.created_at),
    )


def team_event_to_out(event: TeamEvent) -> TeamEventOut:
    return TeamEventOut(
        id=event.id,
        title=event.title,
        eventDate=event.event_date.isoformat(),
        description=event.description,
        owner=event.owner,
        location=event.location,
        createdAt=to_iso(event.created_at),
    )


def ticket_to_out(ticket: Ticket, db: Session) -> TicketOut:
    requester = db.get(User, ticket.requester_id)
    fixed_by = db.get(User, ticket.fixed_by_id) if ticket.fixed_by_id else None
    assignee = db.get(User, ticket.assignee_id) if ticket.assignee_id else None
    evidence_items: list[TicketEvidenceOut] = []
    if ticket.evidence_json:
        try:
            raw_items = json.loads(ticket.evidence_json)
            if isinstance(raw_items, list):
                for raw in raw_items:
                    if not isinstance(raw, dict):
                        continue
                    created_at = str(raw.get("createdAt") or to_iso(datetime.now(timezone.utc)))
                    evidence_items.append(
                        TicketEvidenceOut(
                            id=str(raw.get("id") or str(uuid.uuid4())),
                            text=str(raw.get("text") or ""),
                            imageData=str(raw.get("imageData")) if raw.get("imageData") else None,
                            imageName=str(raw.get("imageName")) if raw.get("imageName") else None,
                            createdAt=created_at,
                        )
                    )
        except json.JSONDecodeError:
            evidence_items = []
    return TicketOut(
        id=ticket.id,
        requesterId=ticket.requester_id,
        requesterEmail=requester.email if requester else "",
        title=ticket.title,
        description=ticket.description,
        category=ticket.category,
        priority=ticket.priority,
        status=ticket.status,
        resolution=ticket.resolution,
        processNotes=ticket.process_notes,
        evidence=evidence_items,
        assigneeId=ticket.assignee_id,
        assigneeEmail=assignee.email if assignee else None,
        slaState=compute_ticket_sla_state(ticket),
        firstResponseDueAt=to_iso(ticket.first_response_due_at) if ticket.first_response_due_at else None,
        resolutionDueAt=to_iso(ticket.resolution_due_at) if ticket.resolution_due_at else None,
        firstRespondedAt=to_iso(ticket.first_responded_at) if ticket.first_responded_at else None,
        resolvedAt=to_iso(ticket.resolved_at) if ticket.resolved_at else None,
        closedAt=to_iso(ticket.closed_at) if ticket.closed_at else None,
        fixedById=ticket.fixed_by_id,
        fixedByEmail=fixed_by.email if fixed_by else None,
        createdAt=to_iso(ticket.created_at),
        updatedAt=to_iso(ticket.updated_at),
    )


def ticket_event_to_out(event: TicketEvent, db: Session) -> TicketEventOut:
    actor_email: str | None = None
    if event.actor_id:
        actor = db.get(User, event.actor_id)
        actor_email = actor.email if actor else None
    payload: dict[str, object] = {}
    if event.payload_json:
        try:
            raw = json.loads(event.payload_json)
            if isinstance(raw, dict):
                payload = raw
        except json.JSONDecodeError:
            payload = {}
    return TicketEventOut(
        id=event.id,
        ticketId=event.ticket_id,
        actorId=event.actor_id,
        actorEmail=actor_email,
        eventType=event.event_type,
        payload=payload,
        createdAt=to_iso(event.created_at),
    )


def maintenance_to_out(record: MaintenanceRecord) -> MaintenanceRecordOut:
    return MaintenanceRecordOut(
        id=record.id,
        maintenanceDate=record.maintenance_date.isoformat(),
        qr=record.qr,
        brand=record.brand,
        model=record.model,
        user=record.user_name,
        serialNumber=record.serial_number,
        consecutive=record.consecutive,
        maintenanceType=record.maintenance_type,
        location=record.location,
        responsibleName=record.responsible_name,
        checks=[
            MaintenanceCheckPayload(
                id=item.id,
                label=item.label,
                category=item.category,
                checked=item.checked,
                observation=item.observation,
            )
            for item in record.checks
        ],
        createdAt=to_iso(record.created_at),
    )


def notification_to_out(notification: Notification) -> NotificationOut:
    return NotificationOut(
        id=notification.id,
        title=notification.title,
        message=notification.message,
        category=notification.category,
        dueDate=notification.due_date.isoformat() if notification.due_date else None,
        read=notification.is_read,
        createdAt=to_iso(notification.created_at),
    )


def article_to_out(article: KnowledgeArticle) -> KnowledgeArticleOut:
    return KnowledgeArticleOut(
        id=article.id,
        title=article.title,
        summary=article.summary,
        content=article.content,
        tags=parse_tags(article.tags),
        createdAt=to_iso(article.created_at),
        updatedAt=to_iso(article.updated_at),
    )


def asset_to_out(asset: Asset) -> AssetOut:
    assigned_user = normalize_assigned_user(asset.user_name)
    return AssetOut(
        id=asset.id,
        assetTag=asset.asset_tag,
        qrCode=asset.qr_code,
        location=asset.location,
        serialNumber=asset.serial_number,
        category=asset.category,
        manufacturer=asset.manufacturer,
        model=asset.model,
        supplier=asset.supplier,
        status=asset.status,
        user=assigned_user,
        condition=asset.condition,
        notes=asset.notes,
        createdAt=to_iso(asset.created_at),
        updatedAt=to_iso(asset.updated_at),
    )


def person_to_out(person: Person) -> PersonOut:
    linked_user_email: str | None = None
    if person.user_id:
        with SessionLocal() as db:
            linked_user = db.get(User, person.user_id)
            linked_user_email = linked_user.email if linked_user else None
    normalized_role = (person.role or "").strip().lower()
    role_value = normalized_role if normalized_role in (UserRole.admin.value, UserRole.developer.value, UserRole.user.value) else UserRole.user.value
    try:
        department_value = normalize_department(person.job_department)
    except HTTPException:
        department_value = DEPARTMENT_OPTIONS[0]
    return PersonOut(
        id=person.id,
        userId=person.user_id,
        userEmail=linked_user_email,
        legacyId=person.legacy_id,
        name=person.name,
        email=person.email,
        title=person.title,
        role=role_value,
        department=department_value,
        mobile=person.mobile,
        notes=person.notes,
        createdAt=to_iso(person.created_at),
        updatedAt=to_iso(person.updated_at),
    )


app = FastAPI(title="WorkPlatform API")
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)


@app.on_event("startup")
def on_startup():
    Base.metadata.create_all(engine)
    with engine.begin() as conn:
        conn.execute(text("ALTER TABLE users ADD COLUMN IF NOT EXISTS preferred_language VARCHAR(5) DEFAULT 'en'"))
        conn.execute(text("ALTER TABLE tickets ADD COLUMN IF NOT EXISTS category VARCHAR(50) DEFAULT 'help'"))
        conn.execute(text("ALTER TABLE tickets ADD COLUMN IF NOT EXISTS process_notes TEXT DEFAULT ''"))
        conn.execute(text("ALTER TABLE tickets ADD COLUMN IF NOT EXISTS evidence_json TEXT DEFAULT '[]'"))
        conn.execute(text("ALTER TABLE tickets ADD COLUMN IF NOT EXISTS assignee_id VARCHAR(36)"))
        conn.execute(text("ALTER TABLE tickets ADD COLUMN IF NOT EXISTS first_response_due_at TIMESTAMPTZ"))
        conn.execute(text("ALTER TABLE tickets ADD COLUMN IF NOT EXISTS resolution_due_at TIMESTAMPTZ"))
        conn.execute(text("ALTER TABLE tickets ADD COLUMN IF NOT EXISTS first_responded_at TIMESTAMPTZ"))
        conn.execute(text("ALTER TABLE tickets ADD COLUMN IF NOT EXISTS resolved_at TIMESTAMPTZ"))
        conn.execute(text("ALTER TABLE tickets ADD COLUMN IF NOT EXISTS closed_at TIMESTAMPTZ"))
        conn.execute(text("CREATE INDEX IF NOT EXISTS ix_tickets_assignee_id ON tickets (assignee_id)"))
        conn.execute(text("UPDATE tickets SET status = 'new' WHERE status = 'open'"))
        conn.execute(text("ALTER TABLE assets ADD COLUMN IF NOT EXISTS qr_code VARCHAR(250) DEFAULT ''"))
        conn.execute(text("ALTER TABLE assets ADD COLUMN IF NOT EXISTS manufacturer VARCHAR(120) DEFAULT ''"))
        conn.execute(text("ALTER TABLE assets ADD COLUMN IF NOT EXISTS model VARCHAR(120) DEFAULT ''"))
        conn.execute(text("ALTER TABLE assets ADD COLUMN IF NOT EXISTS supplier VARCHAR(120) DEFAULT ''"))
        conn.execute(text("ALTER TABLE assets ADD COLUMN IF NOT EXISTS user_name VARCHAR(250) DEFAULT ''"))
        conn.execute(text("ALTER TABLE assets ADD COLUMN IF NOT EXISTS condition VARCHAR(120) DEFAULT ''"))
        conn.execute(text("UPDATE assets SET user_name = 'Unassigned' WHERE btrim(coalesce(user_name, '')) = ''"))
        conn.execute(text("UPDATE assets SET assigned_to = 'Unassigned' WHERE btrim(coalesce(assigned_to, '')) = ''"))
        conn.execute(text("ALTER TABLE people ADD COLUMN IF NOT EXISTS legacy_id INTEGER"))
        conn.execute(text("ALTER TABLE people ADD COLUMN IF NOT EXISTS user_id VARCHAR(36)"))
        conn.execute(text("ALTER TABLE people ADD COLUMN IF NOT EXISTS type VARCHAR(10) DEFAULT 'user'"))
        conn.execute(text("ALTER TABLE people ADD COLUMN IF NOT EXISTS role_id INTEGER DEFAULT 2"))
        conn.execute(text("ALTER TABLE people ADD COLUMN IF NOT EXISTS client_id INTEGER DEFAULT 1"))
        conn.execute(text("ALTER TABLE people ADD COLUMN IF NOT EXISTS name VARCHAR(128) DEFAULT ''"))
        conn.execute(text("ALTER TABLE people ADD COLUMN IF NOT EXISTS email VARCHAR(128) DEFAULT ''"))
        conn.execute(text("ALTER TABLE people ADD COLUMN IF NOT EXISTS title VARCHAR(128) DEFAULT ''"))
        conn.execute(text("ALTER TABLE people ADD COLUMN IF NOT EXISTS role VARCHAR(128) DEFAULT ''"))
        conn.execute(text("ALTER TABLE people ADD COLUMN IF NOT EXISTS job_department VARCHAR(128) DEFAULT ''"))
        conn.execute(text("ALTER TABLE people ADD COLUMN IF NOT EXISTS mobile VARCHAR(100) DEFAULT ''"))
        conn.execute(text("ALTER TABLE people ADD COLUMN IF NOT EXISTS password_hash_legacy VARCHAR(128) DEFAULT ''"))
        conn.execute(text("ALTER TABLE people ADD COLUMN IF NOT EXISTS theme VARCHAR(64) DEFAULT 'skin-blue'"))
        conn.execute(text("ALTER TABLE people ADD COLUMN IF NOT EXISTS sidebar VARCHAR(64) DEFAULT 'opened'"))
        conn.execute(text("ALTER TABLE people ADD COLUMN IF NOT EXISTS layout VARCHAR(64) DEFAULT ''"))
        conn.execute(text("ALTER TABLE people ADD COLUMN IF NOT EXISTS notes TEXT DEFAULT ''"))
        conn.execute(text("ALTER TABLE people ADD COLUMN IF NOT EXISTS signature TEXT DEFAULT ''"))
        conn.execute(text("ALTER TABLE people ADD COLUMN IF NOT EXISTS session_id VARCHAR(255) DEFAULT ''"))
        conn.execute(text("ALTER TABLE people ADD COLUMN IF NOT EXISTS reset_key VARCHAR(255) DEFAULT ''"))
        conn.execute(text("ALTER TABLE people ADD COLUMN IF NOT EXISTS auto_refresh INTEGER DEFAULT 0"))
        conn.execute(text("ALTER TABLE people ADD COLUMN IF NOT EXISTS lang VARCHAR(5) DEFAULT 'en'"))
        conn.execute(text("ALTER TABLE people ADD COLUMN IF NOT EXISTS tickets_notification BOOLEAN DEFAULT FALSE"))
        conn.execute(text("ALTER TABLE people ADD COLUMN IF NOT EXISTS avatar_info TEXT DEFAULT ''"))
        conn.execute(text("ALTER TABLE people ADD COLUMN IF NOT EXISTS created_at TIMESTAMPTZ DEFAULT NOW()"))
        conn.execute(text("ALTER TABLE people ADD COLUMN IF NOT EXISTS updated_at TIMESTAMPTZ DEFAULT NOW()"))
    with SessionLocal() as db:
        existing = db.scalar(select(User).where(User.email == "admin@workplatform.local"))
        if not existing:
            admin = User(
                email="admin@workplatform.local",
                password_hash=hash_password("123456"),
                role=UserRole.admin.value,
                preferred_language="en",
            )
            db.add(admin)
            db.commit()


@app.post("/api/auth/register", response_model=AuthOut)
def register(payload: AuthRegisterIn, _: User = Depends(require_admin), db: Session = Depends(get_db)):
    normalized_email = normalize_login_identity(payload.email)
    existing = db.scalar(select(User).where(User.email == normalized_email))
    if existing:
        raise HTTPException(status_code=409, detail="User already exists")

    user = User(
        email=normalized_email,
        password_hash=hash_password(payload.password),
        role=payload.role.value,
        preferred_language="en",
    )
    db.add(user)
    db.commit()
    db.refresh(user)
    return AuthOut(
        token=create_access_token(user),
        user_email=user.email,
        role=user.role,
        preferred_language=user.preferred_language,
    )


@app.post("/api/auth/login", response_model=AuthOut)
def login(payload: AuthLoginIn, db: Session = Depends(get_db)):
    normalized_email = normalize_login_identity(payload.email)
    user = db.scalar(select(User).where(User.email == normalized_email))
    if not user or not verify_password(payload.password, user.password_hash):
        raise HTTPException(status_code=401, detail="Invalid credentials")
    return AuthOut(
        token=create_access_token(user),
        user_email=user.email,
        role=user.role,
        preferred_language=user.preferred_language,
    )


@app.get("/api/auth/me")
def me(current_user: User = Depends(get_current_user)):
    return {
        "user_email": current_user.email,
        "role": current_user.role,
        "preferred_language": current_user.preferred_language,
    }


@app.patch("/api/account/preferences")
def update_preferences(payload: LanguagePreferenceIn, current_user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    language = payload.preferredLanguage.lower()
    if language not in ("en", "es"):
        raise HTTPException(status_code=400, detail="preferredLanguage must be en or es")
    current_user.preferred_language = language
    db.commit()
    return {"ok": True, "preferred_language": language}


@app.post("/api/auth/change-password")
def change_password(payload: ChangePasswordIn, current_user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    if not verify_password(payload.currentPassword, current_user.password_hash):
        raise HTTPException(status_code=400, detail="Current password is incorrect")
    current_user.password_hash = hash_password(payload.newPassword)
    db.commit()
    return {"ok": True}


@app.get("/api/admin/users", response_model=list[AdminUserOut])
def list_users(_: User = Depends(require_admin), db: Session = Depends(get_db)):
    users = db.scalars(select(User).order_by(User.created_at.desc())).all()
    return [
        AdminUserOut(
            id=user.id,
            email=user.email,
            role=user.role,
            preferredLanguage=user.preferred_language,
            createdAt=to_iso(user.created_at),
        )
        for user in users
    ]


@app.post("/api/admin/users", response_model=AdminUserOut)
def create_user(payload: AdminCreateUserIn, _: User = Depends(require_admin), db: Session = Depends(get_db)):
    normalized_email = normalize_login_identity(payload.email)
    existing = db.scalar(select(User).where(User.email == normalized_email))
    if existing:
        raise HTTPException(status_code=409, detail="User already exists")

    user = User(
        email=normalized_email,
        password_hash=hash_password(payload.password),
        role=payload.role.value,
    )
    db.add(user)
    db.commit()
    db.refresh(user)
    return AdminUserOut(
        id=user.id,
        email=user.email,
        role=user.role,
        preferredLanguage=user.preferred_language,
        createdAt=to_iso(user.created_at),
    )


@app.patch("/api/admin/users/{user_id}", response_model=AdminUserOut)
def update_user(user_id: str, payload: AdminUpdateUserIn, current_admin: User = Depends(require_admin), db: Session = Depends(get_db)):
    user = db.get(User, user_id)
    if not user:
        raise HTTPException(status_code=404, detail="User not found")
    normalized_email = normalize_login_identity(payload.email)
    existing = db.scalar(select(User).where(User.email == normalized_email, User.id != user_id))
    if existing:
        raise HTTPException(status_code=409, detail="Email already in use")
    user.email = normalized_email
    if user.id == current_admin.id and payload.role != UserRole.admin:
        raise HTTPException(status_code=400, detail="Cannot downgrade your own admin role")
    user.role = payload.role.value
    db.commit()
    db.refresh(user)
    return AdminUserOut(
        id=user.id,
        email=user.email,
        role=user.role,
        preferredLanguage=user.preferred_language,
        createdAt=to_iso(user.created_at),
    )


@app.post("/api/admin/users/{user_id}/reset-password")
def reset_user_password(user_id: str, payload: AdminResetPasswordIn, _: User = Depends(require_admin), db: Session = Depends(get_db)):
    user = db.get(User, user_id)
    if not user:
        raise HTTPException(status_code=404, detail="User not found")
    user.password_hash = hash_password(payload.password)
    db.commit()
    return {"ok": True}


@app.delete("/api/admin/users/{user_id}")
def delete_user(user_id: str, current_admin: User = Depends(require_admin), db: Session = Depends(get_db)):
    user = db.get(User, user_id)
    if not user:
        raise HTTPException(status_code=404, detail="User not found")
    if user.id == current_admin.id:
        raise HTTPException(status_code=400, detail="Cannot delete your own admin account")
    db.delete(user)
    db.commit()
    return {"ok": True}


@app.get("/api/admin/people", response_model=list[PersonOut])
def list_people(_: User = Depends(require_admin), db: Session = Depends(get_db)):
    items = db.scalars(select(Person).order_by(Person.name.asc())).all()
    return [person_to_out(item) for item in items]


def generate_placeholder_email(name: str, db: Session) -> str:
    base = "".join(ch for ch in name.lower() if ch.isalnum())
    if not base:
        base = "user"
    candidate = f"{base}@tdcon40.com"
    exists = db.scalar(select(User).where(User.email == candidate))
    if exists:
        raise HTTPException(status_code=409, detail="Generated username already exists for this name; provide email manually")
    return candidate


def resolve_person_user_email(raw_email: str, name: str, db: Session, exclude_user_id: str | None = None) -> str:
    value = (raw_email or "").strip()
    if not value:
        return generate_placeholder_email(name, db)
    normalized = normalize_login_identity(value)
    query = select(User).where(User.email == normalized)
    if exclude_user_id:
        query = query.where(User.id != exclude_user_id)
    existing = db.scalar(query)
    if existing:
        raise HTTPException(status_code=409, detail="User email already exists")
    return normalized


@app.post("/api/admin/people", response_model=PersonOut)
def create_person(payload: PersonIn, _: User = Depends(require_admin), db: Session = Depends(get_db)):
    name = payload.name.strip()
    if not name:
        raise HTTPException(status_code=400, detail="name is required")
    department = normalize_department(payload.department)
    item = Person(
        type="user",
        role_id=2,
        client_id=1,
        name=name,
        email=payload.email.strip().lower(),
        title=payload.title.strip(),
        role=payload.role.value,
        job_department=department,
        mobile=payload.mobile.strip(),
        notes=payload.notes.strip(),
        lang="en",
        tickets_notification=False,
    )
    db.add(item)
    db.flush()
    user_email = resolve_person_user_email(payload.email, name, db)
    linked_user = User(
        email=user_email,
        password_hash=hash_password("12345"),
        role=payload.role.value,
        preferred_language=item.lang if item.lang in ("en", "es") else "en",
    )
    db.add(linked_user)
    db.flush()
    item.user_id = linked_user.id
    db.commit()
    db.refresh(item)
    return person_to_out(item)


@app.patch("/api/admin/people/{person_id}", response_model=PersonOut)
def update_person(person_id: str, payload: PersonIn, _: User = Depends(require_admin), db: Session = Depends(get_db)):
    item = db.get(Person, person_id)
    if not item:
        raise HTTPException(status_code=404, detail="Person not found")
    name = payload.name.strip()
    if not name:
        raise HTTPException(status_code=400, detail="name is required")
    department = normalize_department(payload.department)
    item.name = name
    item.email = payload.email.strip().lower()
    item.title = payload.title.strip()
    item.role = payload.role.value
    item.job_department = department
    item.mobile = payload.mobile.strip()
    item.notes = payload.notes.strip()
    if item.user_id:
        linked_user = db.get(User, item.user_id)
        if linked_user:
            linked_user.role = payload.role.value
            if payload.email.strip():
                linked_user.email = resolve_person_user_email(payload.email, name, db, exclude_user_id=linked_user.id)
    else:
        linked_user = User(
            email=resolve_person_user_email(payload.email, name, db),
            password_hash=hash_password("12345"),
            role=payload.role.value,
            preferred_language=item.lang if item.lang in ("en", "es") else "en",
        )
        db.add(linked_user)
        db.flush()
        item.user_id = linked_user.id
    item.updated_at = datetime.now(timezone.utc)
    db.commit()
    db.refresh(item)
    return person_to_out(item)


@app.delete("/api/admin/people/{person_id}")
def delete_person(person_id: str, _: User = Depends(require_admin), db: Session = Depends(get_db)):
    item = db.get(Person, person_id)
    if not item:
        raise HTTPException(status_code=404, detail="Person not found")
    db.delete(item)
    db.commit()
    return {"ok": True}


@app.get("/api/projects", response_model=list[ProjectOut])
def list_projects(current_user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    projects = db.scalars(select(Project).where(Project.owner_id == current_user.id).order_by(Project.created_at.desc())).all()
    return [project_to_out(item) for item in projects]


@app.post("/api/projects", response_model=ProjectOut)
def create_project(payload: ProjectIn, current_user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    project = Project(owner_id=current_user.id, name=payload.name, description=payload.description)
    db.add(project)
    db.commit()
    db.refresh(project)
    return project_to_out(project)


@app.delete("/api/projects/{project_id}")
def delete_project(project_id: str, current_user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    project = db.scalar(select(Project).where(Project.id == project_id, Project.owner_id == current_user.id))
    if not project:
        raise HTTPException(status_code=404, detail="Project not found")
    db.delete(project)
    db.commit()
    return {"ok": True}


@app.get("/api/tasks", response_model=list[TaskOut])
def list_tasks(current_user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    tasks = db.scalars(select(Task).where(Task.owner_id == current_user.id).order_by(Task.created_at.desc())).all()
    return [task_to_out(item) for item in tasks]


@app.post("/api/tasks", response_model=TaskOut)
def create_task(payload: TaskIn, current_user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    task = Task(
        owner_id=current_user.id,
        title=payload.title,
        details=payload.details,
        project_id=payload.projectId,
        task_date=parse_date(payload.taskDate),
        status="todo",
    )
    db.add(task)
    db.commit()
    db.refresh(task)
    return task_to_out(task)


@app.patch("/api/tasks/{task_id}/status")
def update_task_status(task_id: str, payload: TaskStatusPatch, current_user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    task = db.scalar(select(Task).where(Task.id == task_id, Task.owner_id == current_user.id))
    if not task:
        raise HTTPException(status_code=404, detail="Task not found")
    task.status = payload.status
    db.commit()
    return {"ok": True}


@app.delete("/api/tasks/{task_id}")
def delete_task(task_id: str, current_user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    task = db.scalar(select(Task).where(Task.id == task_id, Task.owner_id == current_user.id))
    if not task:
        raise HTTPException(status_code=404, detail="Task not found")
    db.delete(task)
    db.commit()
    return {"ok": True}


@app.get("/api/notes", response_model=list[NoteOut])
def list_notes(current_user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    notes = db.scalars(select(Note).where(Note.owner_id == current_user.id).order_by(Note.created_at.desc())).all()
    return [note_to_out(item) for item in notes]


@app.post("/api/notes", response_model=NoteOut)
def create_note(payload: NoteIn, current_user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    note = Note(owner_id=current_user.id, title=payload.title, content=payload.content, note_date=parse_date(payload.noteDate))
    db.add(note)
    db.commit()
    db.refresh(note)
    return note_to_out(note)


@app.delete("/api/notes/{note_id}")
def delete_note(note_id: str, current_user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    note = db.scalar(select(Note).where(Note.id == note_id, Note.owner_id == current_user.id))
    if not note:
        raise HTTPException(status_code=404, detail="Note not found")
    db.delete(note)
    db.commit()
    return {"ok": True}


@app.get("/api/notifications", response_model=list[NotificationOut])
def list_notifications(current_user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    notifications = db.scalars(
        select(Notification)
        .where(Notification.owner_id == current_user.id)
        .order_by(Notification.is_read.asc(), Notification.due_date.asc(), Notification.created_at.desc())
    ).all()
    return [notification_to_out(item) for item in notifications]


@app.post("/api/notifications", response_model=NotificationOut)
def create_notification(payload: NotificationIn, current_user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    category = payload.category.strip().lower()
    if category not in ("info", "reminder", "warning"):
        raise HTTPException(status_code=400, detail="category must be info|reminder|warning")
    item = Notification(
        owner_id=current_user.id,
        title=payload.title.strip(),
        message=payload.message.strip(),
        category=category,
        due_date=parse_optional_date(payload.dueDate),
        is_read=False,
    )
    db.add(item)
    db.commit()
    db.refresh(item)
    return notification_to_out(item)


@app.patch("/api/notifications/{notification_id}/read")
def patch_notification_read(
    notification_id: str,
    payload: NotificationReadPatchIn,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    item = db.scalar(select(Notification).where(Notification.id == notification_id, Notification.owner_id == current_user.id))
    if not item:
        raise HTTPException(status_code=404, detail="Notification not found")
    item.is_read = payload.read
    db.commit()
    return {"ok": True}


@app.delete("/api/notifications/{notification_id}")
def delete_notification(notification_id: str, current_user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    item = db.scalar(select(Notification).where(Notification.id == notification_id, Notification.owner_id == current_user.id))
    if not item:
        raise HTTPException(status_code=404, detail="Notification not found")
    db.delete(item)
    db.commit()
    return {"ok": True}


@app.get("/api/knowledge-base", response_model=list[KnowledgeArticleOut])
def list_knowledge_articles(current_user: User = Depends(require_developer_or_admin), db: Session = Depends(get_db)):
    articles = db.scalars(
        select(KnowledgeArticle)
        .order_by(KnowledgeArticle.updated_at.desc())
    ).all()
    return [article_to_out(item) for item in articles]


@app.post("/api/knowledge-base", response_model=KnowledgeArticleOut)
def create_knowledge_article(payload: KnowledgeArticleIn, current_user: User = Depends(require_developer_or_admin), db: Session = Depends(get_db)):
    article = KnowledgeArticle(
        owner_id=current_user.id,
        title=payload.title.strip(),
        summary=payload.summary.strip(),
        content=payload.content.strip(),
        tags=serialize_tags(payload.tags),
    )
    db.add(article)
    db.commit()
    db.refresh(article)
    return article_to_out(article)


@app.patch("/api/knowledge-base/{article_id}", response_model=KnowledgeArticleOut)
def update_knowledge_article(
    article_id: str,
    payload: KnowledgeArticleIn,
    current_user: User = Depends(require_developer_or_admin),
    db: Session = Depends(get_db),
):
    article = db.scalar(select(KnowledgeArticle).where(KnowledgeArticle.id == article_id))
    if not article:
        raise HTTPException(status_code=404, detail="Article not found")
    article.title = payload.title.strip()
    article.summary = payload.summary.strip()
    article.content = payload.content.strip()
    article.tags = serialize_tags(payload.tags)
    article.updated_at = datetime.now(timezone.utc)
    db.commit()
    db.refresh(article)
    return article_to_out(article)


@app.delete("/api/knowledge-base/{article_id}")
def delete_knowledge_article(article_id: str, current_user: User = Depends(require_developer_or_admin), db: Session = Depends(get_db)):
    article = db.scalar(select(KnowledgeArticle).where(KnowledgeArticle.id == article_id))
    if not article:
        raise HTTPException(status_code=404, detail="Article not found")
    db.delete(article)
    db.commit()
    return {"ok": True}


@app.get("/api/assets", response_model=list[AssetOut])
def list_assets(current_user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    if current_user.role in (UserRole.admin.value, UserRole.developer.value):
        items = db.scalars(select(Asset).order_by(Asset.updated_at.desc())).all()
    else:
        items = db.scalars(select(Asset).where(Asset.owner_id == current_user.id).order_by(Asset.updated_at.desc())).all()
    return [asset_to_out(item) for item in items]


@app.post("/api/assets", response_model=AssetOut)
def create_asset(payload: AssetIn, current_user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    status_value = payload.status.strip().lower()
    if status_value not in ("active", "maintenance", "retired", "lost"):
        raise HTTPException(status_code=400, detail="status must be active|maintenance|retired|lost")

    assigned_user = normalize_assigned_user(payload.user)
    qr_class = normalize_qr_class(payload.qrClass)
    generated_tag = build_next_asset_tag(db)
    generated_qr = build_qr_code_from_asset_tag(generated_tag, qr_class)
    item = Asset(
        owner_id=current_user.id,
        asset_tag=generated_tag,
        name=payload.model.strip() or payload.serialNumber.strip().upper() or generated_tag,
        qr_code=generated_qr,
        location=payload.location.strip(),
        serial_number=payload.serialNumber.strip().upper(),
        category=payload.category.strip(),
        manufacturer=payload.manufacturer.strip(),
        model=payload.model.strip(),
        supplier=payload.supplier.strip(),
        status=status_value,
        assigned_to=assigned_user,
        purchase_date=None,
        warranty_until=None,
        notes=payload.notes.strip(),
        user_name=assigned_user,
        condition=payload.condition.strip(),
    )
    db.add(item)
    db.commit()
    db.refresh(item)
    return asset_to_out(item)


@app.patch("/api/assets/{asset_id}", response_model=AssetOut)
def update_asset(asset_id: str, payload: AssetIn, current_user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    if current_user.role in (UserRole.admin.value, UserRole.developer.value):
        item = db.scalar(select(Asset).where(Asset.id == asset_id))
    else:
        item = db.scalar(select(Asset).where(Asset.id == asset_id, Asset.owner_id == current_user.id))
    if not item:
        raise HTTPException(status_code=404, detail="Asset not found")

    status_value = payload.status.strip().lower()
    if status_value not in ("active", "maintenance", "retired", "lost"):
        raise HTTPException(status_code=400, detail="status must be active|maintenance|retired|lost")

    assigned_user = normalize_assigned_user(payload.user)
    qr_class = normalize_qr_class(payload.qrClass)
    qr_code_value = build_qr_code_from_asset_tag(item.asset_tag, qr_class)

    item.name = payload.model.strip() or payload.serialNumber.strip().upper() or item.asset_tag
    item.qr_code = qr_code_value
    item.location = payload.location.strip()
    item.serial_number = payload.serialNumber.strip().upper()
    item.category = payload.category.strip()
    item.manufacturer = payload.manufacturer.strip()
    item.model = payload.model.strip()
    item.supplier = payload.supplier.strip()
    item.status = status_value
    item.assigned_to = assigned_user
    item.purchase_date = None
    item.warranty_until = None
    item.notes = payload.notes.strip()
    item.user_name = assigned_user
    item.condition = payload.condition.strip()
    item.updated_at = datetime.now(timezone.utc)
    db.commit()
    db.refresh(item)
    return asset_to_out(item)


@app.delete("/api/assets/{asset_id}")
def delete_asset(asset_id: str, current_user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    if current_user.role in (UserRole.admin.value, UserRole.developer.value):
        item = db.scalar(select(Asset).where(Asset.id == asset_id))
    else:
        item = db.scalar(select(Asset).where(Asset.id == asset_id, Asset.owner_id == current_user.id))
    if not item:
        raise HTTPException(status_code=404, detail="Asset not found")
    db.delete(item)
    db.commit()
    return {"ok": True}


@app.get("/api/team-events", response_model=list[TeamEventOut])
def list_team_events(current_user: User = Depends(require_developer_or_admin), db: Session = Depends(get_db)):
    events = db.scalars(select(TeamEvent).where(TeamEvent.owner_id == current_user.id).order_by(TeamEvent.created_at.desc())).all()
    return [team_event_to_out(item) for item in events]


@app.post("/api/team-events", response_model=TeamEventOut)
def create_team_event(payload: TeamEventIn, current_user: User = Depends(require_developer_or_admin), db: Session = Depends(get_db)):
    event = TeamEvent(
        owner_id=current_user.id,
        title=payload.title,
        event_date=parse_date(payload.eventDate),
        description=payload.description,
        owner=payload.owner,
        location=payload.location,
    )
    db.add(event)
    db.commit()
    db.refresh(event)
    return team_event_to_out(event)


@app.delete("/api/team-events/{event_id}")
def delete_team_event(event_id: str, current_user: User = Depends(require_developer_or_admin), db: Session = Depends(get_db)):
    event = db.scalar(select(TeamEvent).where(TeamEvent.id == event_id, TeamEvent.owner_id == current_user.id))
    if not event:
        raise HTTPException(status_code=404, detail="Event not found")
    db.delete(event)
    db.commit()
    return {"ok": True}


@app.delete("/api/team-events")
def delete_team_events_by_date(eventDate: str, current_user: User = Depends(require_developer_or_admin), db: Session = Depends(get_db)):
    target_date = parse_date(eventDate)
    events = db.scalars(select(TeamEvent).where(TeamEvent.owner_id == current_user.id, TeamEvent.event_date == target_date)).all()
    for item in events:
        db.delete(item)
    db.commit()
    return {"ok": True}


@app.get("/api/tickets/mine", response_model=list[TicketOut])
def list_my_tickets(current_user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    tickets = db.scalars(
        select(Ticket)
        .where(Ticket.requester_id == current_user.id)
        .order_by(Ticket.created_at.desc())
    ).all()
    return [ticket_to_out(item, db) for item in tickets]


@app.get("/api/tickets/mine/{ticket_id}", response_model=TicketOut)
def get_my_ticket(ticket_id: str, current_user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    ticket = db.scalar(select(Ticket).where(Ticket.id == ticket_id, Ticket.requester_id == current_user.id))
    if not ticket:
        raise HTTPException(status_code=404, detail="Ticket not found")
    return ticket_to_out(ticket, db)


@app.get("/api/tickets/mine/{ticket_id}/events", response_model=list[TicketEventOut])
def list_my_ticket_events(ticket_id: str, current_user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    ticket = db.scalar(select(Ticket).where(Ticket.id == ticket_id, Ticket.requester_id == current_user.id))
    if not ticket:
        raise HTTPException(status_code=404, detail="Ticket not found")
    events = db.scalars(
        select(TicketEvent).where(TicketEvent.ticket_id == ticket_id).order_by(TicketEvent.created_at.asc())
    ).all()
    return [ticket_event_to_out(item, db) for item in events]


@app.delete("/api/tickets/mine/{ticket_id}")
def delete_my_ticket(ticket_id: str, current_user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    ticket = db.scalar(select(Ticket).where(Ticket.id == ticket_id, Ticket.requester_id == current_user.id))
    if not ticket:
        raise HTTPException(status_code=404, detail="Ticket not found")
    events = db.scalars(select(TicketEvent).where(TicketEvent.ticket_id == ticket_id)).all()
    for event in events:
        db.delete(event)
    db.delete(ticket)
    db.commit()
    return {"ok": True}


@app.post("/api/tickets", response_model=TicketOut)
def create_ticket(payload: TicketIn, current_user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    category = normalize_ticket_category(payload.category)
    priority = normalize_ticket_priority(payload.priority)
    now = datetime.now(timezone.utc)
    first_due, resolution_due = calculate_ticket_deadlines(priority, now)
    ticket = Ticket(
        requester_id=current_user.id,
        title=payload.title,
        description=payload.description,
        category=category,
        priority=priority,
        status="new",
        process_notes="",
        evidence_json="[]",
        first_response_due_at=first_due,
        resolution_due_at=resolution_due,
    )
    db.add(ticket)
    db.flush()
    log_ticket_event(
        db,
        ticket,
        current_user.id,
        "created",
        {"status": ticket.status, "priority": priority, "category": category},
    )
    db.commit()
    db.refresh(ticket)
    return ticket_to_out(ticket, db)


@app.get("/api/tickets/open", response_model=list[TicketOut])
def list_open_tickets(_: User = Depends(require_developer_or_admin), db: Session = Depends(get_db)):
    tickets = db.scalars(
        select(Ticket)
        .where(Ticket.status.in_(TICKET_ACTIVE_STATUSES))
        .order_by(Ticket.created_at.asc())
    ).all()
    return [ticket_to_out(item, db) for item in tickets]


@app.get("/api/tickets/open-unassigned", response_model=list[TicketOut])
def list_open_unassigned_tickets(_: User = Depends(require_developer_or_admin), db: Session = Depends(get_db)):
    tickets = db.scalars(
        select(Ticket)
        .where(Ticket.status.in_(TICKET_ACTIVE_STATUSES), Ticket.assignee_id.is_(None))
        .order_by(Ticket.created_at.asc())
    ).all()
    return [ticket_to_out(item, db) for item in tickets]


@app.get("/api/tickets/assigned-mine", response_model=list[TicketOut])
def list_assigned_my_tickets(current_user: User = Depends(require_developer_or_admin), db: Session = Depends(get_db)):
    tickets = db.scalars(
        select(Ticket)
        .where(Ticket.assignee_id == current_user.id)
        .order_by(Ticket.updated_at.desc())
    ).all()
    return [ticket_to_out(item, db) for item in tickets]


@app.get("/api/tickets/assignable-users", response_model=list[TicketAssigneeOut])
def list_ticket_assignable_users(current_user: User = Depends(require_developer_or_admin), db: Session = Depends(get_db)):
    if current_user.role == UserRole.admin.value:
        items = db.scalars(
            select(User)
            .where(User.role.in_([UserRole.admin.value, UserRole.developer.value]))
            .order_by(User.email.asc())
        ).all()
    else:
        items = [current_user]
    return [TicketAssigneeOut(id=item.id, email=item.email, role=item.role) for item in items]


@app.get("/api/tickets/{ticket_id}", response_model=TicketOut)
def get_ticket(ticket_id: str, _: User = Depends(require_developer_or_admin), db: Session = Depends(get_db)):
    ticket = db.get(Ticket, ticket_id)
    if not ticket:
        raise HTTPException(status_code=404, detail="Ticket not found")

    return ticket_to_out(ticket, db)


@app.patch("/api/tickets/{ticket_id}", response_model=TicketOut)
def patch_ticket(ticket_id: str, payload: TicketPatchIn, current_user: User = Depends(require_developer_or_admin), db: Session = Depends(get_db)):
    ticket = db.get(Ticket, ticket_id)
    if not ticket:
        raise HTTPException(status_code=404, detail="Ticket not found")

    next_status = normalize_ticket_status(payload.status) if payload.status is not None else ticket.status
    validate_ticket_transition(ticket.status, next_status)
    previous_status = ticket.status
    ticket.status = next_status
    ticket.resolution = payload.resolution
    if payload.assigneeId is not None:
        requested_assignee_id = payload.assigneeId if payload.assigneeId != "" else None
        validate_assignment_permission(current_user, requested_assignee_id)
        if payload.assigneeId == "":
            ticket.assignee_id = None
        else:
            assignee = db.get(User, payload.assigneeId)
            if not assignee:
                raise HTTPException(status_code=404, detail="Assignee not found")
            if assignee.role not in (UserRole.admin.value, UserRole.developer.value):
                raise HTTPException(status_code=400, detail="Assignee must be admin or developer")
            ticket.assignee_id = assignee.id
    if payload.processNotes is not None:
        ticket.process_notes = payload.processNotes
    if payload.evidence is not None:
        normalized_evidence: list[dict[str, str | None]] = []
        for item in payload.evidence:
            image_data = item.imageData.strip() if item.imageData else None
            if image_data and not image_data.startswith("data:image/"):
                raise HTTPException(status_code=400, detail="evidence imageData must be a data:image/* URL")
            normalized_evidence.append(
                {
                    "id": item.id.strip() or str(uuid.uuid4()),
                    "text": item.text.strip(),
                    "imageData": image_data,
                    "imageName": item.imageName.strip() if item.imageName else None,
                    "createdAt": item.createdAt or to_iso(datetime.now(timezone.utc)),
                }
            )
        ticket.evidence_json = json.dumps(normalized_evidence)
    now = datetime.now(timezone.utc)
    if ticket.status in ("in_progress", "triaged") and ticket.first_responded_at is None:
        ticket.first_responded_at = now
    if ticket.status == "resolved":
        ticket.resolved_at = now
        ticket.closed_at = None
    if ticket.status == "closed":
        ticket.closed_at = now
    if ticket.status == "reopened":
        ticket.resolved_at = None
        ticket.closed_at = None
    ticket.updated_at = datetime.now(timezone.utc)
    ticket.fixed_by_id = current_user.id if ticket.status in ("in_progress", "resolved", "closed") else None
    log_ticket_event(
        db,
        ticket,
        current_user.id,
        "updated",
        {
            "previousStatus": previous_status,
            "status": ticket.status,
            "assigneeId": ticket.assignee_id,
        },
    )
    db.commit()
    db.refresh(ticket)
    return ticket_to_out(ticket, db)


@app.patch("/api/tickets/{ticket_id}/assign", response_model=TicketOut)
def assign_ticket(ticket_id: str, payload: TicketAssignIn, current_user: User = Depends(require_developer_or_admin), db: Session = Depends(get_db)):
    ticket = db.get(Ticket, ticket_id)
    if not ticket:
        raise HTTPException(status_code=404, detail="Ticket not found")
    validate_assignment_permission(current_user, payload.assigneeId)
    if payload.assigneeId is None:
        ticket.assignee_id = None
    else:
        assignee = db.get(User, payload.assigneeId)
        if not assignee:
            raise HTTPException(status_code=404, detail="Assignee not found")
        if assignee.role not in (UserRole.admin.value, UserRole.developer.value):
            raise HTTPException(status_code=400, detail="Assignee must be admin or developer")
        ticket.assignee_id = assignee.id
    ticket.updated_at = datetime.now(timezone.utc)
    log_ticket_event(db, ticket, current_user.id, "assigned", {"assigneeId": ticket.assignee_id})
    db.commit()
    db.refresh(ticket)
    return ticket_to_out(ticket, db)


@app.get("/api/tickets/{ticket_id}/events", response_model=list[TicketEventOut])
def list_ticket_events(ticket_id: str, _: User = Depends(require_developer_or_admin), db: Session = Depends(get_db)):
    ticket = db.get(Ticket, ticket_id)
    if not ticket:
        raise HTTPException(status_code=404, detail="Ticket not found")
    events = db.scalars(
        select(TicketEvent).where(TicketEvent.ticket_id == ticket_id).order_by(TicketEvent.created_at.asc())
    ).all()
    return [ticket_event_to_out(item, db) for item in events]


@app.get("/api/maintenance-records", response_model=list[MaintenanceRecordOut])
def list_maintenance_records(current_user: User = Depends(require_developer_or_admin), db: Session = Depends(get_db)):
    records = db.scalars(
        select(MaintenanceRecord).where(MaintenanceRecord.owner_id == current_user.id).order_by(MaintenanceRecord.created_at.desc())
    ).all()
    return [maintenance_to_out(item) for item in records]


@app.post("/api/maintenance-records", response_model=MaintenanceRecordOut)
def create_maintenance_record(payload: MaintenanceRecordIn, current_user: User = Depends(require_developer_or_admin), db: Session = Depends(get_db)):
    maintenance_type = payload.maintenanceType.upper()
    if maintenance_type not in ("P", "C"):
        raise HTTPException(status_code=400, detail="maintenanceType must be P or C")

    record = MaintenanceRecord(
        owner_id=current_user.id,
        maintenance_date=parse_date(payload.maintenanceDate),
        qr=payload.qr.upper(),
        brand=payload.brand.upper(),
        model=payload.model.upper(),
        user_name=payload.user.upper(),
        serial_number=payload.serialNumber.upper(),
        consecutive=payload.consecutive.upper(),
        maintenance_type=maintenance_type,
        location=payload.location.upper(),
        responsible_name=username_from_email(current_user.email),
    )
    for item in payload.checks:
        record.checks.append(
            MaintenanceCheck(
                id=item.id,
                label=item.label,
                category=item.category,
                checked=item.checked,
                observation=item.observation.upper(),
            )
        )

    db.add(record)
    db.commit()
    db.refresh(record)
    return maintenance_to_out(record)


@app.delete("/api/maintenance-records/{record_id}")
def delete_maintenance_record(record_id: str, current_user: User = Depends(require_developer_or_admin), db: Session = Depends(get_db)):
    record = db.scalar(select(MaintenanceRecord).where(MaintenanceRecord.id == record_id, MaintenanceRecord.owner_id == current_user.id))
    if not record:
        raise HTTPException(status_code=404, detail="Maintenance record not found")
    db.delete(record)
    db.commit()
    return {"ok": True}


CHECK_CELL_MAP = {
    "hardware-general-cleaning": ("N12", "Q12"),
    "hardware-internal-components-cleaning": ("N13", "Q13"),
    "hardware-peripherals-validation": ("N14", "Q14"),
    "hardware-power-system-validation": ("N15", "Q15"),
    "hardware-network-card-validation": ("N16", "Q16"),
    "software-os-diagnosis-correction": ("N19", "Q19"),
    "software-os-driver-updates": ("N20", "Q20"),
    "software-system-files-cleanup": ("N21", "Q21"),
    "software-disk-optimization": ("N22", "Q22"),
    "software-antivirus-check": ("N23", "Q23"),
    "software-virus-definitions-update": ("N24", "Q24"),
    "software-service-pack-installation": ("N25", "Q25"),
    "software-ram-optimization": ("N26", "Q26"),
    "software-disk-capacity-optimization": ("N27", "Q27"),
    "software-authorized-software-policies": ("N28", "Q28"),
}


def _normalize_upper(value: str) -> str:
    return (value or "").strip().upper()


def _sanitize_token(value: str) -> str:
    source = _normalize_upper(value)
    return "".join(ch for ch in source if ch.isalnum() or ch == "-")


def _normalize_consecutive_4(value: str) -> str:
    source = _normalize_upper(value)
    if source.startswith("TDC-"):
        source = source[4:]
    digits = "".join(ch for ch in source if ch.isdigit())
    if not digits:
        return "0000"
    return digits.zfill(4)[-4:]


@app.post("/api/maintenance/export")
async def export_maintenance(
    template: UploadFile = File(...),
    payload: str = Form(...),
    current_user: User = Depends(require_developer_or_admin),
    db: Session = Depends(get_db),
):
    if not template.filename:
        raise HTTPException(status_code=400, detail="Template file is required")

    lower_name = template.filename.lower()
    if not (lower_name.endswith(".xlsx") or lower_name.endswith(".xlsm")):
        raise HTTPException(status_code=400, detail="Template must be an .xlsx or .xlsm file")

    try:
        record = json.loads(payload)
    except json.JSONDecodeError as exc:
        raise HTTPException(status_code=400, detail="Invalid payload JSON") from exc

    required_fields = [
        "id",
        "qr",
        "brand",
        "model",
        "user",
        "serialNumber",
        "consecutive",
        "maintenanceType",
        "maintenanceDate",
        "location",
        "responsibleName",
        "checks",
    ]
    for field in required_fields:
        if field not in record:
            raise HTTPException(status_code=400, detail=f"Missing field: {field}")

    existing_record = db.scalar(
        select(MaintenanceRecord).where(
            MaintenanceRecord.id == str(record["id"]),
            MaintenanceRecord.owner_id == current_user.id,
        )
    )
    if not existing_record:
        raise HTTPException(status_code=404, detail="Maintenance record not found")

    maintenance_type = _normalize_upper(str(record["maintenanceType"]))
    if maintenance_type not in ("P", "C"):
        raise HTTPException(status_code=400, detail="maintenanceType must be P or C")

    brand = _normalize_upper(str(record["brand"]))
    model = _normalize_upper(str(record["model"]))
    qr = _normalize_upper(str(record["qr"]))
    user_name = _normalize_upper(str(record["user"]))
    serial_number = _normalize_upper(str(record["serialNumber"]))
    consecutive = _normalize_upper(str(record["consecutive"]))
    location = _normalize_upper(str(record["location"]))
    responsible_name = _normalize_upper(str(record["responsibleName"]))

    try:
        date_value = datetime.strptime(record["maintenanceDate"], "%Y-%m-%d")
    except ValueError as exc:
        raise HTTPException(status_code=400, detail="maintenanceDate must be YYYY-MM-DD") from exc

    template_bytes = await template.read()
    try:
        workbook = load_workbook(io.BytesIO(template_bytes), keep_vba=lower_name.endswith(".xlsm"))
    except Exception as exc:  # noqa: BLE001
        raise HTTPException(status_code=400, detail="Could not open template workbook") from exc

    sheet = workbook[workbook.sheetnames[0]]
    sheet["L3"] = brand
    sheet["D5"] = qr
    sheet["D6"] = model
    sheet["S6"] = serial_number
    sheet["D7"] = user_name
    sheet["S7"] = date_value
    sheet["J8"] = responsible_name
    sheet["S8"] = location

    for reviewed_cell, observation_cell in CHECK_CELL_MAP.values():
        sheet[reviewed_cell] = ""
        sheet[observation_cell] = ""

    checks = record["checks"] if isinstance(record["checks"], list) else []
    for item in checks:
        check_id = item.get("id")
        mapping = CHECK_CELL_MAP.get(check_id)
        if not mapping:
            continue
        reviewed_cell, observation_cell = mapping
        if bool(item.get("checked")):
            sheet[reviewed_cell] = "X"
        observation = _normalize_upper(str(item.get("observation", "")))
        if observation:
            sheet[observation_cell] = observation

    output = io.BytesIO()
    workbook.save(output)
    output.seek(0)

    file_brand = _sanitize_token(brand) or "NA"
    file_model = _sanitize_token(model) or "NA"
    file_serial = _sanitize_token(serial_number) or "NA"
    file_consecutive = _normalize_consecutive_4(consecutive)
    extension = ".xlsm" if lower_name.endswith(".xlsm") else ".xlsx"
    filename = f"TDC-{file_brand}_{file_model}_{file_serial}_{file_consecutive}{maintenance_type}{extension}"
    encoded_filename = quote(filename)
    content_disposition = f"attachment; filename=\"{filename}\"; filename*=UTF-8''{encoded_filename}"

    media_type = (
        "application/vnd.ms-excel.sheet.macroEnabled.12"
        if extension == ".xlsm"
        else "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

    return StreamingResponse(
        output,
        media_type=media_type,
        headers={"Content-Disposition": content_disposition},
    )
